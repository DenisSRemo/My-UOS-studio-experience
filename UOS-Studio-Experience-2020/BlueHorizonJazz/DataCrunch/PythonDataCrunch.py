import os
import shutil
from openpyxl import load_workbook

class FishingSpot:

       
    Biome = ""
    TimeTravelled = 0
    ChanceAtCatchingFish = 0

class Session:
    def __init__(self):
        self.TimesInBiomes = []
        self.TimesToUnlockBiomes = []
        self.FishingSpots = []
        print("New session created!")
        
    Date = ""
    TimesInBiomes = []
    FishCaught = 0
    SessionLength = ""
    TimesToUnlockBiomes = []
    FishingSpots = []
    SessionID = ""

currentDirectory = os.getcwd()

entries = os.listdir(currentDirectory)

# string is the string to parse over
# wordFromRight is an int of which word you want returned, 0 indexed from the right
def GetRightWordFromString(string, wordFromRight):
    tempString = string.split()
    #print(len(tempString))
    return tempString[len(tempString) - 1 - wordFromRight]


# workbook set-up
workbook = load_workbook(filename = "BHData.xlsx")
print(workbook.sheetnames)
sheets = workbook.sheetnames
sheet = workbook[sheets[0]]
row = 3

def ResetRow():
    row = 3
    while(sheet["A" + str(row)].value != None):
        row = row+1

    print("Latest non-written row: " + str(row))
    return row

row = ResetRow()

# Get the data directory

dataDir = ''
processedDataDir = ''
for entry in entries: # for each thing in current directory
    if(os.path.isdir(os.path.join(currentDirectory, entry))): # if it is a new directory
        potentialDir = os.path.join(currentDirectory, entry)
        print(potentialDir)
        if(os.path.relpath(potentialDir) == 'Data'): # check if the directory is 'Data'
            dataDir = potentialDir
            print(dataDir)
        if(os.path.relpath(potentialDir) == 'ProcessedData'):
            processedDataDir = potentialDir
            print(processedDataDir)


dataFiles = os.listdir(dataDir)

print("######## found " + str(len(dataFiles)) + " files ########")
fileCount = 1;

# read all files in the data directory
for file in dataFiles:
    
    lineStack = []  # create stack so we can FILO and flip the file upside down
    print("File " + str(fileCount))
    #print(file)
    with open(os.path.join(dataDir, file), 'r') as f: # open file
        data = f.readlines()
        #print(data)
        for line in data:
            line.format(0, line.strip()) # strip line of newLines
            lineStack.append(line.rstrip('\n')) # add line to stack ready for processing

    # process data in the stack

    # now, the file is very predictable in its format. from the bottom up we get this:
    
    # 0 - Time in Twilight
    # 1 - Time in Tundra
    # 2 - Time in Swamp
    # 3 - Time in Sunrise
    # 4 - Time in Shallow
    # 5 - Time in Mountain
    # 6 - Time in Normal
    # 7 - Session length

    # this is a given, so we can take advantage of this to get this data.
    lineCount = 0
    fishSpotIndex = 0

    mySession = Session()

    print("")
    
    while(len(lineStack) > 0):
        # pop first item
        tempData = lineStack.pop()

        # first 7 items as stated above
        if(lineCount < 7):
            mySession.TimesInBiomes.append(GetRightWordFromString(tempData, 1))

        # session length
        elif(lineCount == 7):
            mySession.SessionLength = GetRightWordFromString(tempData, 0)

        # fish caught in session
        elif(lineCount == 8):
            if(len(tempData) > 0):
                mySession.FishCaught = GetRightWordFromString(tempData, 0)

        #  date of session
        elif(len(lineStack) == 0):
            mySession.Date = GetRightWordFromString(tempData, 1)
            mySession.SessionID = mySession.Date + GetRightWordFromString(tempData, 0)
            
        # the rest of the data
        else:
            # time to unlock biome, or fishing spot data
            if(len(tempData) < 25):
                if(tempData[0] == 'T'): # time to unlock biome
                    mySession.TimesToUnlockBiomes.append(GetRightWordFromString(tempData, 0))
                else: # fishing chance
                    mySession.FishingSpots.append(FishingSpot())
                    mySession.FishingSpots[fishSpotIndex].ChanceAtCatchingFish = tempData
            else: # fishing spot travel time and biome
                mySession.FishingSpots[fishSpotIndex].TimeTravelled = GetRightWordFromString(tempData, 2)
                mySession.FishingSpots[fishSpotIndex].Biome = GetRightWordFromString(tempData, 0)
                fishSpotIndex = fishSpotIndex + 1
        lineCount = lineCount + 1

    # Print data and export into excel
    sheet = workbook[sheets[0]]
    sheet["A" + str(row)].value = mySession.SessionID

    print("Data in session class:")
    print("")
    print("Session " + mySession.Date)
    sheet["B" + str(row)].value = mySession.Date
    print("-------------------------------------")
    print("Times in Biomes")
    biomeCount = 0
    for item in mySession.TimesInBiomes:
        sheet[chr(99+biomeCount)+str(row)].value = str(item)
        print("    " + str(item))
        biomeCount = biomeCount + 1

    print("-------------------------------------")
    print("Time in session: " + mySession.SessionLength)
    sheet["J"+str(row)].value = mySession.SessionLength

    print("-------------------------------------")
    print("Fish caught: " + str(mySession.FishCaught))
    sheet["K"+str(row)].value = str(mySession.FishCaught)

    print("-------------------------------------")
    print("Times to unlock biomes")
    for time in mySession.TimesToUnlockBiomes:
        print("    " + str(time))

    print("-------------------------------------")
    print("Fishing Chances")
    for spot in mySession.FishingSpots:
        sheet = workbook[sheets[1]]
        print(sheet)
        row = ResetRow()
        print("    "  + spot.Biome)
        print("    "  + str(spot.TimeTravelled))
        print("    "  + str(spot.ChanceAtCatchingFish))

        sheet["A"+str(row)].value = '=A'+str(row-1)+'+1'
        sheet["B"+str(row)].value = mySession.SessionID
        sheet["C"+str(row)].value = spot.Biome
        sheet["D"+str(row)].value = spot.TimeTravelled
        sheet["E"+str(row)].value = spot.ChanceAtCatchingFish
        print("-------")
    
    print("")
    print("")
    print("")
    print("")

    fileCount = fileCount + 1
    sheet = workbook[sheets[0]]
    row = ResetRow()

    # move file to processed
    shutil.move(os.path.join(dataDir, file), os.path.join(processedDataDir, file))
    #break

workbook.save(filename = "BHData.xlsx")

#input()
